import requests
import json
import jsonpath
import openpyxl

# API url
url = "http://thetestingworldapi.com/api/studentsDetails"

# Read file content
file = open("C:/Kiran Kumar SY/Python/venvScripts1/venvScripts/com/python/UdemyRestApi/DataDrivenTesting/JsonTestData.json",'r')
fileinput = file.read()

# Convert to json format
json_request = json.loads(fileinput)

# Read xl file
wb = openpyxl.load_workbook("C:/Kiran Kumar SY/Python/venvScripts1/venvScripts/com/python/UdemyRestApi/DataDrivenTesting/JsonTestData.xlsx")
sh = wb['Sheet1']
rowcnt = sh.max_row
print(rowcnt)

for i in range(2,rowcnt+1):
    cellfirst_name = sh.cell(row=i,column=1)
    cellmiddle_name = sh.cell(row=i,column=2)
    celllast_name = sh.cell(row=i,column=3)
    celldate_of_birth = sh.cell(row=i,column=4)

    json_request['first_name'] = cellfirst_name.value
    json_request['middle_name'] = cellmiddle_name.value
    json_request['last_name'] = celllast_name
    json_request['date_of_birth'] = celldate_of_birth

    response = requests.post(url,json_request)
    print(response.text)
    print(response.status_code)

    assert response.status_code==201
