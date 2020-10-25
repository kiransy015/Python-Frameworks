import json
import jsonpath
import requests
import openpyxl

class Demo:

    def __init__(self,filepath,sheetname):
        global wb
        global sh
        wb = openpyxl.load_workbook(filepath)
        sh = wb[sheetname]

    def getRowCnt(self):
        rowcnt = sh.max_row
        return rowcnt

    def getColumnCnt(self):
        colcnt = sh.max_column
        return colcnt

    def getKeyList(self):
        c = sh.max_column
        li=[]
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)
        return li

    def updatejsonrequest(self,rowno,jsonrequest,sheetname,KeyList):
        sh = wb[sheetname]
        colcnt = sh.max_column

        for i in range(1,colcnt+1):
            cell = sh.cell(row=rowno,column=i)
            jsonrequest[KeyList[i-1]]=cell.value
        return jsonrequest